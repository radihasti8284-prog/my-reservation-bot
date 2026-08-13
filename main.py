from flask import Flask, request, jsonify, send_from_directory, send_file
import os
import requests
import sqlite3
import json
import time
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import threading
import jdatetime
import traceback

# ====== کتابخانه‌های Cloudinary ======
import cloudinary
import cloudinary.uploader
import cloudinary.api

# ====== کتابخانه‌های PDF ======
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from io import BytesIO

# ====== کتابخانه‌های فارسی در PDF ======
import arabic_reshaper
from bidi.algorithm import get_display

# ============================================================
#   راه‌اندازی اولیه
# ============================================================
app = Flask(__name__)

TOKEN = "8971000707:AAESYFI--ALKEXQgDN7c0yb9SjEBbQQN3BM"
UPLOAD_FOLDER = 'static/uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}

ADMIN_IDS = [int(x.strip()) for x in os.environ.get("ADMIN_IDS", "").split(",") if x.strip().isdigit()]
print(f"👑 Admin IDs: {ADMIN_IDS}")

# ====== تنظیمات Cloudinary (با دیباگ) ======
CLOUDINARY_CLOUD_NAME = os.environ.get("CLOUDINARY_CLOUD_NAME")
CLOUDINARY_API_KEY = os.environ.get("CLOUDINARY_API_KEY")
CLOUDINARY_API_SECRET = os.environ.get("CLOUDINARY_API_SECRET")

if CLOUDINARY_CLOUD_NAME and CLOUDINARY_API_KEY and CLOUDINARY_API_SECRET:
    cloudinary.config(
        cloud_name=CLOUDINARY_CLOUD_NAME,
        api_key=CLOUDINARY_API_KEY,
        api_secret=CLOUDINARY_API_SECRET
    )
    print("✅ Cloudinary configured successfully.")
else:
    print("⚠️ Cloudinary credentials not found. Using local upload fallback.")


# ============================================================
#   تابع آپلود (با دیباگ قوی و پشتیبانی از آپلود محلی)
# ============================================================
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/api/upload', methods=['POST'])
def upload_file():
    try:
        # بررسی وجود فایل
        if 'file' not in request.files:
            return jsonify({"status": "error", "message": "هیچ فایلی ارسال نشده است."}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"status": "error", "message": "هیچ فایلی انتخاب نشده است."}), 400

        # بررسی فرمت فایل
        if not allowed_file(file.filename):
            return jsonify({
                "status": "error",
                "message": "فرمت فایل پشتیبانی نمی‌شود. فقط PNG, JPG, JPEG, GIF, WEBP, PDF"
            }), 400

        # ====== تلاش برای آپلود در Cloudinary ======
        if CLOUDINARY_CLOUD_NAME and CLOUDINARY_API_KEY and CLOUDINARY_API_SECRET:
            try:
                upload_result = cloudinary.uploader.upload(
                    file,
                    folder="m4cut_receipts",
                    transformation={"quality": "auto", "fetch_format": "auto"},
                    timeout=30
                )
                image_url = upload_result['secure_url']
                print(f"✅ File uploaded to Cloudinary: {image_url}")
                return jsonify({"status": "ok", "url": image_url})
            except Exception as e:
                print(f"⚠️ Cloudinary upload failed: {e}")
                # اگر Cloudinary خطا داد، به آپلود محلی برگرد
                print("🔄 Falling back to local upload...")

        # ====== آپلود محلی (Fallback) ======
        try:
            filename = f"{int(time.time())}_{secure_filename(file.filename)}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            local_url = f"/static/uploads/{filename}"
            print(f"✅ File uploaded locally: {local_url}")
            return jsonify({"status": "ok", "url": local_url})
        except Exception as e:
            print(f"❌ Local upload failed: {e}")
            return jsonify({"status": "error", "message": f"خطا در ذخیره فایل: {str(e)}"}), 500

    except Exception as e:
        print(f"❌ upload_file critical error: {e}")
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"خطای سرور: {str(e)}"}), 500)


# ============================================================
#   توابع کمکی PDF (تبدیل فارسی)
# ============================================================
def reshape_persian(text):
    """تبدیل متن فارسی به شکل صحیح برای نمایش در PDF"""
    if not text:
        return ""
    try:
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    except:
        return text


def generate_appointment_pdf(appointments, title="گزارش نوبت‌ها"):
    """تولید PDF از لیست نوبت‌ها با پشتیبانی از فارسی"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm,
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()

    # استایل فارسی
    persian_style = ParagraphStyle(
        'PersianStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        alignment=TA_RIGHT,
        encoding='utf-8'
    )

    elements = []

    # عنوان
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    elements.append(Paragraph(reshape_persian(title), title_style))

    # تاریخ امروز (شمسی)
    now = jdatetime.datetime.now()
    date_str = now.strftime("%Y/%m/%d - %H:%M")
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        alignment=TA_RIGHT,
        textColor=colors.gray
    )
    elements.append(Paragraph(reshape_persian(f"تاریخ: {date_str}"), date_style))
    elements.append(Spacer(1, 20))

    # جدول داده‌ها
    if appointments and len(appointments) > 0:
        # هدر جدول
        headers = ['ردیف', 'نام کاربر', 'شماره تماس', 'خدمت', 'تاریخ', 'ساعت', 'وضعیت']
        data = [headers]

        for idx, app in enumerate(appointments, 1):
            status_map = {
                'pending': 'در انتظار',
                'confirmed': 'تأیید شده',
                'completed': 'انجام شده',
                'cancelled': 'لغو شده'
            }
            row = [
                str(idx),
                app.get('user_name', 'نامشخص'),
                app.get('phone', '---'),
                app.get('service_name', '---'),
                app.get('appointment_date', '---'),
                app.get('appointment_time', '---'),
                status_map.get(app.get('status', ''), app.get('status', '---'))
            ]
            data.append(row)

        # ساخت جدول
        table = Table(data, colWidths=[30, 80, 80, 100, 80, 60, 80])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.gold),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)

        # تعداد کل
        elements.append(Spacer(1, 10))
        count_style = ParagraphStyle(
            'CountStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            alignment=TA_RIGHT
        )
        elements.append(Paragraph(reshape_persian(f"تعداد کل نوبت‌ها: {len(appointments)}"), count_style))
    else:
        elements.append(Paragraph(reshape_persian("هیچ نوبتی یافت نشد."), persian_style))

    # فوتر
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.gray
    )
    elements.append(Paragraph(reshape_persian("M4Cut © 2026 - گزارش خودکار"), footer_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ============================================================
#   دیتابیس
# ============================================================
def get_db():
    try:
        conn = sqlite3.connect('data.db')
        conn.row_factory = sqlite3.Row
        return conn
    except Exception as e:
        print(f"❌ Database connection error: {e}")
        traceback.print_exc()
        return None


def init_db():
    try:
        conn = get_db()
        if conn is None:
            print("❌ Failed to connect to database for init")
            return
        cursor = conn.cursor()

        cursor.execute('''CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            telegram_id INTEGER UNIQUE,
            name TEXT,
            phone TEXT,
            is_admin INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS services (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            duration INTEGER DEFAULT 30,
            price INTEGER DEFAULT 0,
            description TEXT,
            is_active INTEGER DEFAULT 1
        )''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS appointments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            service_id INTEGER,
            appointment_date TEXT,
            appointment_time TEXT,
            status TEXT DEFAULT 'pending',
            receipt TEXT,
            notification_sent INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (service_id) REFERENCES services(id)
        )''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS work_schedule (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            day_of_week INTEGER NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL,
            capacity INTEGER DEFAULT 2,
            is_active INTEGER DEFAULT 1
        )''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS daily_capacity (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            appointment_date TEXT NOT NULL,
            appointment_time TEXT NOT NULL,
            booked_count INTEGER DEFAULT 0,
            max_capacity INTEGER DEFAULT 2,
            UNIQUE(appointment_date, appointment_time)
        )''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            key TEXT UNIQUE NOT NULL,
            value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS broadcasts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            message TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')

        try:
            cursor.execute("PRAGMA table_info(appointments)")
            cols = [c[1] for c in cursor.fetchall()]
            if 'notification_sent' not in cols:
                cursor.execute("ALTER TABLE appointments ADD COLUMN notification_sent INTEGER DEFAULT 0")
        except Exception as e:
            print(f"⚠️ Error altering appointments: {e}")

        try:
            cursor.execute("PRAGMA table_info(services)")
            cols = [c[1] for c in cursor.fetchall()]
            if 'description' not in cols:
                cursor.execute("ALTER TABLE services ADD COLUMN description TEXT")
        except Exception as e:
            print(f"⚠️ Error altering services: {e}")

        cursor.execute("SELECT COUNT(*) FROM settings WHERE key = 'support_contact'")
        if cursor.fetchone()[0] == 0:
            cursor.execute("INSERT INTO settings (key, value) VALUES ('support_contact', '@Tvpnred')")

        cursor.execute("SELECT COUNT(*) FROM services")
        if cursor.fetchone()[0] == 0:
            services = [('✂️ کوتاهی مو', 30, 200000, '✨ کوتاهی و اصلاح مو با جدیدترین متدها')]
            cursor.executemany("INSERT INTO services (name, duration, price, description) VALUES (?, ?, ?, ?)",
                               services)

        cursor.execute("SELECT COUNT(*) FROM work_schedule")
        if cursor.fetchone()[0] == 0:
            default = [
                (0, '09:00', '18:00', 2, 1),
                (1, '09:00', '18:00', 2, 1),
                (2, '09:00', '18:00', 2, 1),
                (3, '09:00', '18:00', 2, 1),
                (4, '09:00', '18:00', 2, 1),
                (5, '09:00', '14:00', 1, 0),
                (6, '09:00', '14:00', 0, 0)
            ]
            cursor.executemany(
                "INSERT INTO work_schedule (day_of_week, start_time, end_time, capacity, is_active) VALUES (?, ?, ?, ?, ?)",
                default
            )

        conn.commit()
        conn.close()
        print("✅ Database initialized successfully.")
    except Exception as e:
        print(f"❌ Critical init_db error: {e}")
        traceback.print_exc()


init_db()


# ============================================================
#   توابع کمکی اصلی
# ============================================================
def send_message(chat_id, text, reply_markup=None):
    try:
        url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
        payload = {"chat_id": chat_id, "text": text}
        if reply_markup:
            payload["reply_markup"] = json.dumps(reply_markup)
        response = requests.post(url, json=payload, timeout=10)
        if response.status_code != 200:
            print(f"⚠️ Telegram API error: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"⚠️ Error sending message: {e}")


def get_or_create_user(telegram_id, name, phone):
    if not telegram_id:
        return None
    try:
        conn = get_db()
        if conn is None:
            return None
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM users WHERE telegram_id = ?", (telegram_id,))
        user = cursor.fetchone()
        if user:
            conn.close()
            return dict(user)
        is_admin = 1 if telegram_id in ADMIN_IDS else 0
        cursor.execute(
            "INSERT INTO users (telegram_id, name, phone, is_admin) VALUES (?, ?, ?, ?)",
            (telegram_id, name, phone, is_admin)
        )
        conn.commit()
        cursor.execute("SELECT * FROM users WHERE telegram_id = ?", (telegram_id,))
        user = cursor.fetchone()
        conn.close()
        return dict(user)
    except Exception as e:
        print(f"❌ get_or_create_user error: {e}")
        return None


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_all_users():
    try:
        conn = get_db()
        if conn is None:
            return []
        cursor = conn.cursor()
        cursor.execute("SELECT telegram_id FROM users")
        users = cursor.fetchall()
        conn.close()
        return [u['telegram_id'] for u in users]
    except Exception as e:
        print(f"❌ get_all_users error: {e}")
        return []


def send_broadcast(message):
    users = get_all_users()
    success = 0
    for chat_id in users:
        try:
            send_message(chat_id, f"📢 اعلان همگانی:\n\n{message}")
            success += 1
            time.sleep(0.1)
        except Exception:
            pass
    return success


def get_support_contact():
    try:
        conn = get_db()
        if conn is None:
            return '@Tvpnred'
        cursor = conn.cursor()
        cursor.execute("SELECT value FROM settings WHERE key = 'support_contact'")
        res = cursor.fetchone()
        conn.close()
        return res['value'] if res else '@Tvpnred'
    except Exception as e:
        print(f"❌ get_support_contact error: {e}")
        return '@Tvpnred'


def update_support_contact(val):
    try:
        conn = get_db()
        if conn is None:
            return
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE settings SET value = ?, updated_at = CURRENT_TIMESTAMP WHERE key = 'support_contact'",
            (val,)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"❌ update_support_contact error: {e}")


def get_daily_capacity(date, time_slot):
    try:
        conn = get_db()
        if conn is None:
            return 2
        cursor = conn.cursor()
        cursor.execute(
            "SELECT booked_count, max_capacity FROM daily_capacity WHERE appointment_date = ? AND appointment_time = ?",
            (date, time_slot)
        )
        res = cursor.fetchone()
        conn.close()
        if res:
            return res['max_capacity'] - res['booked_count']
        return 2
    except Exception as e:
        print(f"❌ get_daily_capacity error: {e}")
        return 2


def increment_capacity(date, time_slot):
    try:
        conn = get_db()
        if conn is None:
            return
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO daily_capacity (appointment_date, appointment_time, booked_count, max_capacity) "
            "VALUES (?, ?, 1, 2) ON CONFLICT(appointment_date, appointment_time) DO UPDATE SET booked_count = booked_count + 1",
            (date, time_slot)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"❌ increment_capacity error: {e}")


def decrement_capacity(date, time_slot):
    try:
        conn = get_db()
        if conn is None:
            return
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE daily_capacity SET booked_count = booked_count - 1 "
            "WHERE appointment_date = ? AND appointment_time = ? AND booked_count > 0",
            (date, time_slot)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"❌ decrement_capacity error: {e}")


# ============================================================
#   Scheduler یادآوری
# ============================================================
def reminder_job():
    while True:
        try:
            conn = get_db()
            if conn is None:
                time.sleep(60)
                continue
            cursor = conn.cursor()
            today = datetime.now().strftime("%Y/%m/%d")
            tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y/%m/%d")

            for date_target in [tomorrow, today]:
                cursor.execute('''
                    SELECT a.id, a.user_id, a.appointment_date, a.appointment_time, 
                           u.telegram_id, s.name as service_name
                    FROM appointments a
                    JOIN users u ON a.user_id = u.id
                    JOIN services s ON a.service_id = s.id
                    WHERE a.appointment_date = ? AND a.status = 'confirmed' AND a.notification_sent = 0
                ''', (date_target,))
                apps = cursor.fetchall()

                for app in apps:
                    msg = (f"🔔 یادآوری نوبت{' فردا' if date_target == tomorrow else ' امروز'}!\n\n"
                           f"📅 {app['appointment_date']}\n"
                           f"🕐 {app['appointment_time']}\n"
                           f"💇 {app['service_name']}\n\n"
                           "⚠️ لطفاً ۱۰ دقیقه قبل حضور داشته باشید.")
                    send_message(app['telegram_id'], msg)
                    cursor.execute("UPDATE appointments SET notification_sent = 1 WHERE id = ?", (app['id'],))

            conn.commit()
            conn.close()
        except Exception as e:
            print(f"⚠️ Scheduler error: {e}")
            traceback.print_exc()
        time.sleep(3600)


try:
    threading.Thread(target=reminder_job, daemon=True).start()
    print("✅ Scheduler started.")
except Exception as e:
    print(f"❌ Failed to start scheduler: {e}")


# ============================================================
#   وب‌هوک تلگرام
# ============================================================
@app.route('/webhook', methods=['POST'])
def webhook():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "Invalid JSON"}), 400

        print(f"📩 Webhook received")

        if 'message' in data:
            msg = data['message']
            chat_id = msg['chat']['id']
            text = msg.get('text', '')
            user = msg.get('from', {})
            telegram_id = user.get('id')
            full_name = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()

            get_or_create_user(telegram_id, full_name, "")

            if text == '/start':
                send_message(chat_id,
                             "✨ سلام دوست عزیز به **M4Cut** خوش آمدید! ✨\n\n"
                             "💇‍♂️ **آرایشگاه تخصصی مردانه**\n"
                             "با ما بهترین تجربه‌ی کوتاهی مو را داشته باشید.\n\n"
                             "🔹 **خدمات ما:**\n"
                             "✂️ کوتاهی موی حرفه‌ای (مخصوص آقایان)\n"
                             "✨ اصلاح و استایل‌دهی با جدیدترین متدها\n"
                             "💆‍♂️ مشاوره رایگان قبل از هر سرویس\n\n"
                             "💳 **بیعانه:** ۲۰۰,۰۰۰ تومان\n"
                             "⚠️ در صورت عدم حضور، نوبت از بین می‌رود.\n\n"
                             f"📞 **پشتیبانی:** {get_support_contact()}\n\n"
                             "👇 برای رزرو کلیک کنید:",
                             reply_markup={
                                 "inline_keyboard": [[
                                     {"text": "📅 رزرو نوبت",
                                      "web_app": {"url": "https://my-reservation-bot.onrender.com/static/landing.html"}}
                                 ]]
                             }
                             )
            elif text == '/admin':
                conn = get_db()
                if conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT is_admin FROM users WHERE telegram_id = ?", (chat_id,))
                    row = cursor.fetchone()
                    conn.close()
                    if row and row['is_admin'] == 1:
                        send_message(chat_id,
                                     "👋 پنل ادمین",
                                     reply_markup={
                                         "inline_keyboard": [[
                                             {"text": "📊 مدیریت", "web_app": {
                                                 "url": "https://my-reservation-bot.onrender.com/static/admin.html"}}
                                         ]]
                                     }
                                     )
                    else:
                        send_message(chat_id, "⛔ دسترسی ندارید.")
                else:
                    send_message(chat_id, "❌ خطا در ارتباط با دیتابیس.")
            else:
                send_message(chat_id, "❓ از /start استفاده کنید.")
        return {"status": "ok"}, 200
    except Exception as e:
        print(f"❌ Webhook error: {e}")
        traceback.print_exc()
        return {"status": "error", "message": str(e)}, 500


# ============================================================
#   API های عمومی
# ============================================================
@app.route('/api/auth', methods=['POST'])
def auth_user():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "Invalid data"}), 400
        user = get_or_create_user(data.get('telegram_id'), data.get('name'), data.get('phone'))
        if user:
            return jsonify({"status": "ok", "user": user})
        return jsonify({"status": "error", "message": "Failed to create user"}), 500
    except Exception as e:
        print(f"❌ auth_user error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/services', methods=['GET'])
def get_services():
    try:
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database error"}), 500
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM services WHERE is_active = 1")
        services = cursor.fetchall()
        conn.close()
        return jsonify({"status": "ok", "services": [dict(s) for s in services]})
    except Exception as e:
        print(f"❌ get_services error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({"status": "error", "message": "No file"}), 400
        file = request.files['file']
        if file.filename == '':
            return jsonify({"status": "error", "message": "No file selected"}), 400
        if file and allowed_file(file.filename):
            # آپلود به Cloudinary
            upload_result = cloudinary.uploader.upload(
                file,
                folder="m4cut_receipts",
                transformation={"quality": "auto", "fetch_format": "auto"}
            )
            image_url = upload_result['secure_url']
            return jsonify({"status": "ok", "url": image_url})
        return jsonify({"status": "error", "message": "Invalid format"}), 400
    except Exception as e:
        print(f"❌ upload_file error: {e}")
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/appointments', methods=['POST'])
def create_appointment():
    try:
        data = request.get_json()
        telegram_id = data.get('telegram_id')
        service_id = data.get('service_id')
        app_date = data.get('date')
        app_time = data.get('time')
        receipt_url = data.get('receipt')

        if get_daily_capacity(app_date, app_time) <= 0:
            return jsonify({"status": "error", "message": "ظرفیت این ساعت تکمیل شده است."}), 400

        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database error"}), 500
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM users WHERE telegram_id = ?", (telegram_id,))
        user = cursor.fetchone()
        if not user:
            user_dict = get_or_create_user(telegram_id, "کاربر", "")
            if user_dict:
                cursor.execute("SELECT id FROM users WHERE telegram_id = ?", (telegram_id,))
                user = cursor.fetchone()
        if not user:
            conn.close()
            return jsonify({"status": "error", "message": "User not found."}), 404

        cursor.execute('''
            INSERT INTO appointments (user_id, service_id, appointment_date, appointment_time, status, receipt)
            VALUES (?, ?, ?, ?, 'pending', ?)
        ''', (user['id'], service_id, app_date, app_time, receipt_url))
        conn.commit()
        conn.close()
        increment_capacity(app_date, app_time)
        return jsonify({"status": "ok", "message": "نوبت ثبت شد! منتظر تأیید ادمین باشید."})
    except Exception as e:
        print(f"❌ create_appointment error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/appointments/user/<int:telegram_id>', methods=['GET'])
def get_user_appointments(telegram_id):
    try:
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database error"}), 500
        cursor = conn.cursor()
        cursor.execute('''
            SELECT a.*, s.name as service_name
            FROM appointments a
            JOIN users u ON a.user_id = u.id
            JOIN services s ON a.service_id = s.id
            WHERE u.telegram_id = ?
            ORDER BY a.appointment_date DESC
        ''', (telegram_id,))
        apps = cursor.fetchall()
        conn.close()
        return jsonify({"status": "ok", "appointments": [dict(a) for a in apps]})
    except Exception as e:
        print(f"❌ get_user_appointments error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/appointments/<int:appointment_id>', methods=['PUT'])
def update_appointment(appointment_id):
    try:
        data = request.get_json()
        new_date = data.get('date')
        new_time = data.get('time')

        if get_daily_capacity(new_date, new_time) <= 0:
            return jsonify({"status": "error", "message": "ظرفیت این ساعت تکمیل شده است."}), 400

        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database error"}), 500
        cursor = conn.cursor()
        cursor.execute("SELECT appointment_date, appointment_time FROM appointments WHERE id = ?", (appointment_id,))
        old = cursor.fetchone()
        if not old:
            conn.close()
            return jsonify({"status": "error", "message": "نوبت پیدا نشد."}), 404

        cursor.execute('''
            UPDATE appointments 
            SET appointment_date = ?, appointment_time = ? 
            WHERE id = ? AND status = 'pending'
        ''', (new_date, new_time, appointment_id))
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({"status": "error", "message": "فقط نوبت‌های در انتظار قابل ویرایش هستند."}), 400

        decrement_capacity(old['appointment_date'], old['appointment_time'])
        increment_capacity(new_date, new_time)
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "message": "نوبت ویرایش شد."})
    except Exception as e:
        print(f"❌ update_appointment error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/appointments/<int:appointment_id>/cancel', methods=['POST'])
def cancel_appointment(appointment_id):
    try:
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database error"}), 500
        cursor = conn.cursor()
        cursor.execute("SELECT appointment_date, appointment_time FROM appointments WHERE id = ?", (appointment_id,))
        app = cursor.fetchone()
        cursor.execute("UPDATE appointments SET status = 'cancelled' WHERE id = ?", (appointment_id,))
        conn.commit()
        conn.close()
        if app:
            decrement_capacity(app['appointment_date'], app['appointment_time'])
        return jsonify({"status": "ok", "message": "نوبت لغو شد."})
    except Exception as e:
        print(f"❌ cancel_appointment error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ============================================================
#   API های ادمین (با دیباگ قوی)
# ============================================================
@app.route('/api/admin/appointments', methods=['GET'])
def admin_get_appointments():
    try:
        status = request.args.get('status')
        date = request.args.get('date')
        user = request.args.get('user')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        offset = (page - 1) * limit

        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()
        query = '''
            SELECT a.*, u.name as user_name, u.phone, u.telegram_id, s.name as service_name
            FROM appointments a
            JOIN users u ON a.user_id = u.id
            JOIN services s ON a.service_id = s.id
            WHERE 1=1
        '''
        params = []
        if status:
            query += " AND a.status = ?"
            params.append(status)
        if date:
            query += " AND a.appointment_date = ?"
            params.append(date)
        if user:
            query += " AND u.telegram_id = ?"
            params.append(user)

        count_query = query.replace(
            'SELECT a.*, u.name as user_name, u.phone, u.telegram_id, s.name as service_name',
            'SELECT COUNT(*) as total'
        )
        cursor.execute(count_query, params)
        total = cursor.fetchone()['total']

        query += " ORDER BY a.appointment_date DESC LIMIT ? OFFSET ?"
        params.extend([limit, offset])
        cursor.execute(query, params)
        apps = cursor.fetchall()
        conn.close()

        return jsonify({
            "status": "ok",
            "appointments": [dict(a) for a in apps],
            "pagination": {
                "total": total,
                "page": page,
                "limit": limit,
                "pages": (total + limit - 1) // limit if limit > 0 else 1
            }
        })
    except Exception as e:
        print(f"❌ admin_get_appointments error: {e}")
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/admin/appointments/<int:appointment_id>/status', methods=['POST'])
def admin_update_status(appointment_id):
    try:
        data = request.get_json()
        new_status = data.get('status')

        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()
        cursor.execute('''
            SELECT a.user_id, u.telegram_id, a.appointment_date, a.appointment_time, s.name as service_name
            FROM appointments a
            JOIN users u ON a.user_id = u.id
            JOIN services s ON a.service_id = s.id
            WHERE a.id = ?
        ''', (appointment_id,))
        app = cursor.fetchone()

        cursor.execute("UPDATE appointments SET status = ?, notification_sent = 0 WHERE id = ?",
                       (new_status, appointment_id))
        conn.commit()
        conn.close()

        if app and new_status == 'confirmed':
            msg = (f"✅ نوبت شما تأیید شد!\n"
                   f"📅 {app['appointment_date']}\n"
                   f"🕐 {app['appointment_time']}\n"
                   f"💇 {app['service_name']}\n\n"
                   "⚠️ در صورت عدم حضور، نوبت از بین می‌رود.")
            send_message(app['telegram_id'], msg)

        return jsonify({"status": "ok", "message": f"وضعیت به {new_status} تغییر کرد."})
    except Exception as e:
        print(f"❌ admin_update_status error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/admin/users', methods=['GET'])
def admin_get_users():
    try:
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()
        cursor.execute("SELECT id, telegram_id, name, phone, is_admin, created_at FROM users ORDER BY created_at DESC")
        users = cursor.fetchall()
        conn.close()
        return jsonify({"status": "ok", "users": [dict(u) for u in users]})
    except Exception as e:
        print(f"❌ admin_get_users error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/admin/users/<int:telegram_id>', methods=['PUT'])
def admin_update_user(telegram_id):
    try:
        data = request.get_json()
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()
        if data.get('name'):
            cursor.execute("UPDATE users SET name = ? WHERE telegram_id = ?", (data['name'], telegram_id))
        cursor.execute("UPDATE users SET is_admin = ? WHERE telegram_id = ?", (data.get('is_admin', 0), telegram_id))
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "message": "کاربر به‌روزرسانی شد."})
    except Exception as e:
        print(f"❌ admin_update_user error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/admin/users/<int:telegram_id>', methods=['DELETE'])
def admin_delete_user(telegram_id):
    try:
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()
        cursor.execute("DELETE FROM users WHERE telegram_id = ?", (telegram_id,))
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "message": "کاربر حذف شد."})
    except Exception as e:
        print(f"❌ admin_delete_user error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/admin/stats', methods=['GET'])
def admin_get_stats():
    try:
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM users")
        total_users = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM appointments")
        total_appointments = cursor.fetchone()[0]
        cursor.execute("SELECT status, COUNT(*) as count FROM appointments GROUP BY status")
        status_counts = cursor.fetchall()
        status_stats = {row['status']: row['count'] for row in status_counts}
        today = datetime.now().strftime("%Y/%m/%d")
        cursor.execute("SELECT COUNT(*) FROM appointments WHERE appointment_date = ?", (today,))
        today_appointments = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM appointments WHERE appointment_date >= ?", (today,))
        upcoming_appointments = cursor.fetchone()[0]
        conn.close()
        return jsonify({
            "status": "ok",
            "stats": {
                "total_users": total_users,
                "total_appointments": total_appointments,
                "today_appointments": today_appointments,
                "upcoming_appointments": upcoming_appointments,
                "status": {
                    "pending": status_stats.get('pending', 0),
                    "confirmed": status_stats.get('confirmed', 0),
                    "completed": status_stats.get('completed', 0),
                    "cancelled": status_stats.get('cancelled', 0)
                }
            }
        })
    except Exception as e:
        print(f"❌ admin_get_stats error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/admin/support', methods=['GET'])
def get_support():
    try:
        return jsonify({"status": "ok", "support_contact": get_support_contact()})
    except Exception as e:
        print(f"❌ get_support error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/admin/support', methods=['POST'])
def update_support():
    try:
        data = request.get_json()
        val = data.get('support_contact')
        if not val:
            return jsonify({"status": "error", "message": "مقدار الزامی است."}), 400
        update_support_contact(val)
        return jsonify({"status": "ok", "message": "شماره پشتیبانی به‌روزرسانی شد."})
    except Exception as e:
        print(f"❌ update_support error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/admin/schedule', methods=['GET'])
def get_schedule():
    try:
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM work_schedule ORDER BY day_of_week")
        schedule = cursor.fetchall()
        conn.close()
        return jsonify({"status": "ok", "schedule": [dict(s) for s in schedule]})
    except Exception as e:
        print(f"❌ get_schedule error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/admin/schedule', methods=['POST'])
def update_schedule():
    try:
        data = request.get_json()
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE work_schedule 
            SET start_time = ?, end_time = ?, capacity = ?, is_active = ? 
            WHERE day_of_week = ?
        ''', (data['start_time'], data['end_time'], data.get('capacity', 2), data.get('is_active', 1),
              data['day_of_week']))
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "message": "ساعت کاری به‌روز شد."})
    except Exception as e:
        print(f"❌ update_schedule error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/admin/broadcast', methods=['POST'])
def admin_broadcast():
    try:
        data = request.get_json()
        message = data.get('message')
        if not message:
            return jsonify({"status": "error", "message": "Message required"}), 400
        count = send_broadcast(message)
        conn = get_db()
        if conn:
            cursor = conn.cursor()
            cursor.execute("INSERT INTO broadcasts (message) VALUES (?)", (message,))
            conn.commit()
            conn.close()
        return jsonify({"status": "ok", "message": f"پیام به {count} کاربر ارسال شد."})
    except Exception as e:
        print(f"❌ admin_broadcast error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ====== API جدید: گزارش روزانه PDF ======
@app.route('/api/admin/daily-report/pdf', methods=['GET'])
def daily_report_pdf():
    try:
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()

        today = datetime.now().strftime("%Y/%m/%d")

        cursor.execute('''
            SELECT a.*, u.name as user_name, u.phone, s.name as service_name
            FROM appointments a
            JOIN users u ON a.user_id = u.id
            JOIN services s ON a.service_id = s.id
            WHERE a.appointment_date = ?
            ORDER BY a.appointment_time ASC
        ''', (today,))
        apps = cursor.fetchall()
        conn.close()

        appointments = [dict(a) for a in apps]

        # عنوان شامل تاریخ شمسی
        now = jdatetime.datetime.now()
        persian_date = now.strftime("%Y/%m/%d")
        title = f"گزارش روزانه مشتریان - {persian_date}"

        pdf_buffer = generate_appointment_pdf(appointments, title)

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"daily-report-{now.strftime('%Y-%m-%d')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        print(f"❌ daily_report_pdf error: {e}")
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


# ====== API جدید: گزارش کامل PDF ======
@app.route('/api/admin/full-report/pdf', methods=['GET'])
def full_report_pdf():
    try:
        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()

        cursor.execute('''
            SELECT a.*, u.name as user_name, u.phone, s.name as service_name
            FROM appointments a
            JOIN users u ON a.user_id = u.id
            JOIN services s ON a.service_id = s.id
            ORDER BY a.appointment_date DESC, a.appointment_time ASC
        ''')
        apps = cursor.fetchall()
        conn.close()

        appointments = [dict(a) for a in apps]
        title = "گزارش کامل نوبت‌ها"

        pdf_buffer = generate_appointment_pdf(appointments, title)

        return send_file(
            pdf_buffer,
            as_attachment=True,
            download_name=f"full-report-{datetime.now().strftime('%Y-%m-%d')}.pdf",
            mimetype='application/pdf'
        )
    except Exception as e:
        print(f"❌ full_report_pdf error: {e}")
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


# ====== API قبلی اکسل (برای موارد خاص) ======
@app.route('/api/admin/export/excel', methods=['GET'])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from io import BytesIO

        conn = get_db()
        if conn is None:
            return jsonify({"status": "error", "message": "Database connection failed"}), 500
        cursor = conn.cursor()
        cursor.execute('''
            SELECT a.*, u.name as user_name, u.phone, s.name as service_name
            FROM appointments a
            JOIN users u ON a.user_id = u.id
            JOIN services s ON a.service_id = s.id
            ORDER BY a.appointment_date DESC
        ''')
        apps = cursor.fetchall()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "نوبت‌ها"

        headers = ['شناسه', 'کاربر', 'شماره تماس', 'خدمت', 'تاریخ', 'ساعت', 'وضعیت', 'تاریخ ثبت']
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for row_idx, app in enumerate(apps, 2):
            ws.cell(row=row_idx, column=1, value=app['id'])
            ws.cell(row=row_idx, column=2, value=app['user_name'])
            ws.cell(row=row_idx, column=3, value=app['phone'])
            ws.cell(row=row_idx, column=4, value=app['service_name'])
            ws.cell(row=row_idx, column=5, value=app['appointment_date'])
            ws.cell(row=row_idx, column=6, value=app['appointment_time'])
            ws.cell(row=row_idx, column=7, value=app['status'])
            ws.cell(row=row_idx, column=8, value=app['created_at'])

        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 18

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"nobat-ha-{datetime.now().strftime('%Y%m%d')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"❌ export_excel error: {e}")
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/persian_date', methods=['GET'])
def get_persian_date():
    try:
        now = jdatetime.datetime.now()
        return jsonify({
            "status": "ok",
            "date": now.strftime("%Y/%m/%d"),
            "time": now.strftime("%H:%M"),
            "weekday": now.strftime("%A")
        })
    except Exception as e:
        print(f"❌ get_persian_date error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ============================================================
#   سرو فایل‌های استاتیک
# ============================================================
@app.route('/static/<path:path>')
def serve_static(path):
    try:
        return send_from_directory('static', path)
    except Exception as e:
        print(f"❌ serve_static error: {e}")
        return jsonify({"status": "error", "message": "File not found"}), 404


@app.route('/')
def home():
    return "✅ M4Cut روشن است! (نسخه نهایی با Cloudinary و PDF)"


# ============================================================
#   اجرا
# ============================================================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    print(f"🚀 Starting server on port {port}...")
    app.run(host="0.0.0.0", port=port, debug=False)